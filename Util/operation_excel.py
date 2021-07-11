#!/usr/bin/env python
# -*- encoding: utf-8 -*-

import openpyxl
import sys
import os
import json

from Util.operation_json import OperationJson
from Util.operation_init import HandleInit


base_path = os.getcwd()
sys.path.append(base_path)

file_path1 = 'D:/uatotest/Case/Testcase.xlsx'



class HandExcel():

    def __init__(self,file_name=None):
        if file_name:
            self.file_name = file_name
        else:
            self.file_name = file_path1
        self.open_excel = openpyxl.load_workbook(self.file_name)
        self.data =self.get_sheet_data()


    #加载所有的sheet内容
    def get_sheet_data(self,index=None):
        sheet_name = self.open_excel.sheetnames
        if index == None:
            index = 0
        data = self.open_excel[sheet_name[index]]
        return data

    #获取某一个单元格内容
    def get_cell_value(self,row,cols):
        data = self.data.cell(row=row+1,column=cols+1).value
        return data


    #获取行数
    def get_rows(self):
        row = self.data.max_row
        return row


    #获取某一行的内容
    def get_rows_value(self,row):
        row_list = []
        for i in self.data[row]:
            row_list.append(i.value)
        return row_list

    #获取某一列的内容
    def get_columns_value(self,key=None):
        columns_list = []
        if key==None:
            key='A'
        columns_list_data = self.data[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list


    #获取行号
    def get_rows_number(self,case_id):
        num=1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num=num+1
        return num


    #获取excel所有的数据
    def get_excel_data(self):
        data_list =[]
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i+2))
        return data_list


    #获取请求数据关键字key
    def get_request_data(self,row):
        hi = HandleInit()
        col = hi.get_reque_data()
        request_key_data = self.get_cell_value(row,col)
        return request_key_data


    #通过获取关键字key拿到json数据
    def get_data_for_json(self,row):
        opera_json = OperationJson()
        request_data = opera_json.get_data(self.get_request_data(row))
        return request_data



    #写入数据
    def excel_write_data(self, row, cols, value):
        wb = self.open_excel
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(file_path1)



    #拆分单元格数据
    def split_data_one(self,row,col):
        # imooc_005>data:banner:id
        data = self.get_cell_value(row, col)
        rule_key = data.split(">")[0]
        rule_data = data.split(">")[1]
        return rule_key, rule_data

    #拆分数据
    def split_data_two(self,data):
        # imooc_005>data:banner:id
        rule_key = data.split(">")[0]
        rule_data = data.split(">")[1]
        return rule_key, rule_data



if __name__ == "__main__":
    handle = HandExcel()
    data = handle.get_rows()
    print(type(data))
    print(data)
    data2  = handle.split_data_two('api3/getcourseintro>errorDesc')
    print(data2)
    print(type(data2))


